import customtkinter as ctk
from tkinter import messagebox as msg
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import json
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader, PdfWriter

# --- CONFIGURAÇÕES GLOBAIS ---
ARQUIVO_DB = "banco_dados.xlsx"
CONFIG_LAYOUT = "layout_config.json"

class PDFEngine:
    """Motor de geração de PDF baseado em coordenadas dinâmicas."""
    
    @staticmethod
    def carregar_coordenadas():
        if not os.path.exists(CONFIG_LAYOUT):
            return None
        with open(CONFIG_LAYOUT, "r") as f:
            return json.load(f)

    @staticmethod
    def gerar_recibo(cliente, itens, total, output_path, modelo_pdf, campos_extras=None):
        """
        Gera o PDF. 
        'campos_extras' deve ser um dicionário: {"nome_no_json": "valor_digitado"}
        """
        coords = PDFEngine.carregar_coordenadas()
        if not coords:
            msg.showerror("Erro", "Layout não calibrado! Rode o picker.py primeiro.")
            return False

        tmp_path = f"temp_{datetime.now().strftime('%H%M%S')}.pdf"
        c = canvas.Canvas(tmp_path, pagesize=A4)

        # 1. CAMPOS PADRÃO (Sempre existem)
        if "cliente" in coords:
            c.drawString(coords["cliente"][0], coords["cliente"][1], str(cliente))
        
        if "data" in coords:
            c.drawString(coords["data"][0], coords["data"][1], datetime.now().strftime("%d/%m/%Y"))

        # ==========================================================
        # 2. ADICIONAR NOVOS CAMPOS FIXOS AQUI (Ex: Paciente, CNPJ)
        # ==========================================================
        # Se você adicionou 'paciente' no picker.py, ele aparecerá no coords.
        if campos_extras:
            for chave, valor in campos_extras.items():
                if chave in coords:
                    x, y = coords[chave]
                    c.drawString(x, y, str(valor))
        # ==========================================================

        # 3. TABELA DE ITENS
        if "tabela_inicio" in coords:
            x, y = coords["tabela_inicio"]
            for desc, ref, qtd, val_unit, subtotal in itens:
                # Se houver referência (ex: paciente), concatena na descrição
                texto_item = f"{desc} ({ref})" if ref and ref != "—" else desc
                c.drawString(x, y, texto_item)
                
                # Colunas fixas (ajuste os números somados a 'x' se necessário)
                c.drawString(x + 250, y, str(qtd))
                c.drawString(x + 320, y, f"R${val_unit:.2f}")
                c.drawString(x + 420, y, f"R${subtotal:.2f}")
                y -= 20 

        # 4. TOTAL
        if "total" in coords:
            c.drawString(coords["total"][0], coords["total"][1], f"R${total:.2f}")

        c.save()

        # Mesclagem com o modelo original
        try:
            reader = PdfReader(modelo_pdf)
            overlay = PdfReader(tmp_path)
            page = reader.pages[0]
            page.merge_page(overlay.pages[0])
            writer = PdfWriter()
            writer.add_page(page)
            with open(output_path, "wb") as f_out:
                writer.write(f_out)
            return True
        except Exception as e:
            msg.showerror("Erro PDF", f"Falha ao mesclar PDF: {e}")
            return False
        finally:
            if os.path.exists(tmp_path): os.remove(tmp_path)

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Gerador de Recibos Profissional")
        self.geometry("450x420")
        self._inicializar_db()
        self._criar_widgets()

    def _inicializar_db(self):
        """Cria o Excel se não existir."""
        if not os.path.exists(ARQUIVO_DB):
            wb = Workbook()
            ws1 = wb.active; ws1.title = "Clientes"; ws1.append(["ID", "Nome"])
            wb.create_sheet("Itens").append(["ID", "Descrição", "Valor"])
            wb.create_sheet("Registros").append(["Cliente", "Data", "Referencia", "Item", "Qtd", "Unit", "Total"])
            wb.save(ARQUIVO_DB)

    def _criar_widgets(self):
        ctk.CTkLabel(self, text="📊 Gestão de Recibos", font=("Arial", 22, "bold")).pack(pady=20)
        
        btn_style = {"width": 280, "height": 45, "corner_radius": 8}
        ctk.CTkButton(self, text="➕ Cadastrar Novo Cliente", command=self.janela_cliente, **btn_style).pack(pady=10)
        ctk.CTkButton(self, text="📦 Gerenciar Itens/Serviços", command=self.janela_item, **btn_style).pack(pady=10)
        ctk.CTkButton(self, text="📝 Emitir Novo Recibo", command=self.janela_emissao, **btn_style, fg_color="#2c3e50").pack(pady=10)
        
        status = "🟢 Layout Configurado" if os.path.exists(CONFIG_LAYOUT) else "🔴 Layout Não Encontrado"
        ctk.CTkLabel(self, text=status, font=("Arial", 11)).pack(side="bottom", pady=15)

    def janela_cliente(self):
        # Janela simples para salvar nome no Excel
        top = ctk.CTkToplevel(self); top.title("Novo Cliente"); top.geometry("300x180"); top.attributes("-topmost", True)
        ctk.CTkLabel(top, text="Nome do Cliente:").pack(pady=5)
        ent = ctk.CTkEntry(top, width=220); ent.pack(pady=5)
        def salvar():
            if not ent.get(): return
            wb = load_workbook(ARQUIVO_DB); wb["Clientes"].append([wb["Clientes"].max_row, ent.get()])
            wb.save(ARQUIVO_DB); top.destroy(); msg.showinfo("Sucesso", "Cliente salvo!")
        ctk.CTkButton(top, text="Salvar", command=salvar).pack(pady=15)

    def janela_item(self):
        # Janela para salvar serviços e preços no Excel
        top = ctk.CTkToplevel(self); top.title("Novo Item"); top.geometry("300x230"); top.attributes("-topmost", True)
        ctk.CTkLabel(top, text="Descrição:").pack(); e_d = ctk.CTkEntry(top); e_d.pack(pady=5)
        ctk.CTkLabel(top, text="Preço Unitário:").pack(); e_p = ctk.CTkEntry(top); e_p.pack(pady=5)
        def salvar():
            try:
                wb = load_workbook(ARQUIVO_DB); wb["Itens"].append([wb["Itens"].max_row, e_d.get(), float(e_p.get().replace(",","."))])
                wb.save(ARQUIVO_DB); top.destroy(); msg.showinfo("Sucesso", "Item salvo!")
            except: msg.showerror("Erro", "Valor inválido!")
        ctk.CTkButton(top, text="Salvar", command=salvar).pack(pady=10)

    def janela_emissao(self):
        wb = load_workbook(ARQUIVO_DB)
        clientes = [row[1] for row in wb["Clientes"].iter_rows(min_row=2, values_only=True)]
        itens_db = {row[1]: row[2] for row in wb["Itens"].iter_rows(min_row=2, values_only=True)}

        if not clientes: return msg.showwarning("Atenção", "Cadastre clientes primeiro!")

        top = ctk.CTkToplevel(self); top.title("Emissão de Recibo"); top.geometry("650x650"); top.attributes("-topmost", True)

        # SELEÇÃO DE CLIENTE
        ctk.CTkLabel(top, text="Cliente:").pack()
        c_var = ctk.StringVar(value=clientes[0])
        ctk.CTkOptionMenu(top, values=clientes, variable=c_var).pack(pady=5)

        # ==========================================================
        # 1. ADICIONAR INPUT DE CAMPO EXTRA AQUI (Ex: Paciente)
        # ==========================================================
        ctk.CTkLabel(top, text="Informação Extra (Ex: Nome do Paciente / OS):").pack()
        e_extra = ctk.CTkEntry(top, width=350, placeholder_text="Digite aqui...")
        e_extra.pack(pady=5)
        # ==========================================================

        # ADICIONAR ITENS AO CARRINHO
        frame_item = ctk.CTkFrame(top); frame_item.pack(pady=10, padx=20, fill="x")
        i_var = ctk.StringVar(value=list(itens_db.keys())[0] if itens_db else "")
        ctk.CTkOptionMenu(frame_item, values=list(itens_db.keys()), variable=i_var).grid(row=0, column=0, px=5)
        e_qtd = ctk.CTkEntry(frame_item, width=60); e_qtd.insert(0, "1"); e_qtd.grid(row=0, column=1, px=5)
        
        carrinho = []
        txt = ctk.CTkTextbox(top, height=150, width=550); txt.pack(pady=10)

        def add():
            nome = i_var.get(); qtd = float(e_qtd.get()); preco = itens_db[nome]; sub = qtd * preco
            carrinho.append((nome, "—", qtd, preco, sub)) # O "—" pode ser a referência por linha
            txt.insert("end", f"{nome} | {qtd}x | R${sub:.2f}\n")
        
        ctk.CTkButton(frame_item, text="Adicionar", command=add, width=100).grid(row=0, column=2, px=5)

        def finalizar():
            if not carrinho: return
            total = sum(c[4] for c in carrinho)
            
            # Pegamos o valor do campo extra
            dados_extras = {"paciente": e_extra.get()} # A chave 'paciente' deve bater com o picker.py

            # Caminhos
            modelo = os.path.join(os.getcwd(), "IDV", "modelo_recibo.pdf")
            pasta_saida = os.path.join(os.getcwd(), "recibos")
            os.makedirs(pasta_saida, exist_ok=True)
            nome_arq = f"Recibo_{c_var.get()}_{datetime.now().strftime('%H%M%S')}.pdf"
            caminho_final = os.path.join(pasta_saida, nome_arq)

            sucesso = PDFEngine.gerar_recibo(c_var.get(), carrinho, total, caminho_final, modelo, campos_extras=dados_extras)
            if sucesso:
                msg.showinfo("Sucesso", f"PDF Gerado: {nome_arq}")
                top.destroy()

        ctk.CTkButton(top, text="🚀 GERAR RECIBO PDF", command=finalizar, height=50, fg_color="green").pack(pady=20)

if __name__ == "__main__":
    app = App()
    app.mainloop()